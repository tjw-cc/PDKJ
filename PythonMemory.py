
import xlrd2
from PyQt5.QtGui import QBrush, QColor
from PyQt5.QtWidgets import QTableWidgetItem
import openpyxl as op


# 对excel表的读取
class PythonMemory(object):
    def __init__(self, parent=None):
        super(PythonMemory, self).__init__(parent)

    # 读取excel表格
    def read_excel(path):
        # print(path)
        # 打开excel文件
        data1 = xlrd2.open_workbook(path)

        # 打开指定的sheet页
        table = data1.sheets()[0]
        print(table.nrows)
        # 创建一个空列表，存储Excel的数据
        tables = []
        for rown in range(3,table.nrows):
            array = {'RealU': '', 'StandardU': '', 'Deviation': '', "Temperature":'', "Balance":''}
            if(table.cell_value(rown, 0) is not None):
                array['RealU'] = table.cell_value(rown, 0)
                array['StandardU'] = table.cell_value(rown, 1)
                array['Deviation'] = table.cell_value(rown, 0)-table.cell_value(rown, 1)
                array['Temperature'] = table.cell_value(rown, 3)
                array['Balance'] = table.cell_value(rown, 4)

            tables.append(array)
        return tables


    # 将tableWidget的数据写入excel表格中
    def write_qt_excel(tableWidget, path,m):

        # 获取tableWidget的行列数
        row = tableWidget.rowCount()
        cols = tableWidget.columnCount()
        # print(row)
        # print(cols)
        print(m)
        # 打开MY_EXCEL.xlsx文件
        wb = op.load_workbook(path)
        # 激活工作区
        ws = wb.active
        for i in range(row):
            # tableWidget.insertRow(row)
            for j in range(cols):
                try:
                    # 计算实际电压和基准电压的差值
                    if (j== 3):
                        n = int(ws.cell(row=i + 4, column=1).value)-int(ws.cell(row=i + 4, column=2).value)
                        a = abs(n)
                        # print(str(n))
                        newItem = QTableWidgetItem(str(abs(n)))
                        # 如果差值大于阈值，则将差值字体设置为红色
                        try:
                            if (a>=int(m)):
                                newItem.setForeground(QBrush(QColor(255, 0, 0)))
                        except:
                            print("阈值为0")
                        # print("\033[31m这是红色字体\033[0m")
                        # print(str(rowslist[j]))
                        tableWidget.setItem(i, j-1, newItem)
                        ws.cell(row=i+4, column=j+1).value = str(tableWidget.item(i, j).text())

                    # print(tableWidget.)
                    if (tableWidget.item(i, j) != None):
                        # print(tableWidget.item(i, j).text())
                        ws.cell(row=i+4, column=j+1).value = str(tableWidget.item(i, j).text())
                    else:
                        ws.cell(row=i + 4, column=j + 1).value = ''
                except:
                    print('保存失败')


        wb.save(path)  # 保存excel表
        print('保存数据')







    # 从excel表格内读取数据导入tableWidget中
    def read_excel_qt(tableWidget, path):
        # 打开文件
        workbook = xlrd2.open_workbook(path)
        # 获取所有sheet
        sheet2_name = workbook.sheet_names()[0]
        # 根据sheet索引或者名称获取sheet内容
        sheet1 = workbook.sheet_by_index(0)  # sheet索引从0开始
        cols = sheet1.col_values(0)  # 获取第1列内容
        print(cols)
        # 获取整行和整列的值（数组）
        for i in range(len(cols)-3):
            # 从第三行开始获取excel每行内容
            rowslist = sheet1.row_values(i+3)

            # 在tablewidget中添加行
            row = tableWidget.rowCount()
            # print(row)
            # 插入行后写入数据
            # tableWidget.insertRow(row)
            for j in range(len(rowslist)):

                # 把数据写入tablewidget中
                newItem = QTableWidgetItem(str(rowslist[j]))
                # print(str(rowslist[j]))
                tableWidget.setItem(i, j, newItem)



    # 将tablewidget中的数据删除
    def cleartableWidget(tableWidget):

        row = tableWidget.rowCount()
        for i in range(row):

            # tableWidget.insertRow(row)
            for j in range(5):
                tableWidget.setItem(i, j, None)
        print("已清除")


